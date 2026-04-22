#!/usr/bin/env python3
"""Fill blank Tokyo workbook rows using Tokyo-only unique corporate matches.

This updates the uploaded workbook in-place with a narrower matching rule:
- Only inspect Sheet1 rows whose 登録番号 is between 20834 and 7640 inclusive
- Only consider rows where both 法人番号 and 本所住所 are blank
- Match the legal name against the official Tokyo corporate dataset
- Fill only when the name maps to exactly one corporation within Tokyo
- Leave not-found or Tokyo-duplicate names blank
"""

from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = REPO_ROOT / "data/东京都 ).xlsx"
TOKYO_CSV_PATHS = [
    REPO_ROOT / "data/raw/13_tokyo_all_20260331_01.csv",
    REPO_ROOT / "data/raw/13_tokyo_all_20260331_02.csv",
]

MIN_REGISTRATION_NUMBER = 7640
MAX_REGISTRATION_NUMBER = 20834

NTA_CORPORATE_NUMBER_COL = 1
NTA_NAME_COL = 6
NTA_PREF_COL = 9
NTA_CITY_COL = 10
NTA_ADDRESS_COL = 11
NTA_POSTCODE_COL = 15


def normalize_name(text: str) -> str:
    text = unicodedata.normalize("NFKC", text or "")
    return "".join(ch for ch in text if not ch.isspace())


def strip_trailing_alias(name: str) -> str:
    return re.sub(r"[\(（].*?[\)）]\s*$", "", name or "").strip()


def format_postcode(postcode: str) -> str:
    digits = re.sub(r"\D", "", postcode or "")
    if len(digits) == 7:
        return f"〒{digits[:3]}-{digits[3:]}"
    return ""


def format_address(postcode: str, prefecture: str, city: str, address: str) -> str:
    postal = format_postcode(postcode)
    body = f"{prefecture or ''}{city or ''}{address or ''}".strip()
    return f"{postal}{body}" if postal and body else body


def get_sheet_indexes(headers: list[object]) -> dict[str, int]:
    required = ["登録番号", "会社名", "法人番号", "本所住所"]
    indexes: dict[str, int] = {}
    for key in required:
        if key not in headers:
            raise KeyError(f"Missing expected column: {key}")
        indexes[key] = headers.index(key) + 1
    return indexes


def collect_blank_targets(ws, indexes: dict[str, int]) -> dict[str, list[int]]:
    targets: dict[str, list[int]] = {}
    for row_idx in range(2, ws.max_row + 1):
        registration_value = ws.cell(row_idx, indexes["登録番号"]).value
        try:
            registration_number = int(registration_value)
        except (TypeError, ValueError):
            continue

        if not (MIN_REGISTRATION_NUMBER <= registration_number <= MAX_REGISTRATION_NUMBER):
            continue

        corporate_number = ws.cell(row_idx, indexes["法人番号"]).value
        address = ws.cell(row_idx, indexes["本所住所"]).value
        if corporate_number or address:
            continue

        company_name = ws.cell(row_idx, indexes["会社名"]).value
        search_key = normalize_name(strip_trailing_alias(company_name))
        if not search_key:
            continue
        targets.setdefault(search_key, []).append(row_idx)
    return targets


def scan_tokyo_candidates(search_keys: set[str]) -> dict[str, list[dict[str, str]]]:
    candidates_by_key: dict[str, dict[str, dict[str, str]]] = {}

    for path in TOKYO_CSV_PATHS:
        if not path.exists():
            raise FileNotFoundError(f"Tokyo CSV not found: {path}")

        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for row in reader:
                if len(row) <= NTA_POSTCODE_COL:
                    continue
                if row[NTA_PREF_COL] != "東京都":
                    continue

                search_key = normalize_name(row[NTA_NAME_COL])
                if search_key not in search_keys:
                    continue

                corporate_number = row[NTA_CORPORATE_NUMBER_COL]
                candidates_by_key.setdefault(search_key, {})
                candidates_by_key[search_key][corporate_number] = {
                    "corporate_number": corporate_number,
                    "address": format_address(
                        row[NTA_POSTCODE_COL],
                        row[NTA_PREF_COL],
                        row[NTA_CITY_COL],
                        row[NTA_ADDRESS_COL],
                    ),
                }

    return {
        key: list(corporate_number_map.values())
        for key, corporate_number_map in candidates_by_key.items()
    }


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    workbook = load_workbook(WORKBOOK_PATH)
    worksheet = workbook["Sheet1"]
    headers = [worksheet.cell(1, column_idx).value for column_idx in range(1, worksheet.max_column + 1)]
    indexes = get_sheet_indexes(headers)

    targets = collect_blank_targets(worksheet, indexes)
    candidates_by_key = scan_tokyo_candidates(set(targets))

    updated_rows = 0
    remaining_blank_duplicate_rows = 0
    remaining_blank_not_found_rows = 0

    for search_key, row_indexes in targets.items():
        candidates = candidates_by_key.get(search_key, [])
        if len(candidates) == 1:
            candidate = candidates[0]
            for row_idx in row_indexes:
                worksheet.cell(row_idx, indexes["法人番号"]).value = candidate["corporate_number"]
                worksheet.cell(row_idx, indexes["本所住所"]).value = candidate["address"]
                updated_rows += 1
        elif len(candidates) == 0:
            remaining_blank_not_found_rows += len(row_indexes)
        else:
            remaining_blank_duplicate_rows += len(row_indexes)

    workbook.save(WORKBOOK_PATH)

    print(f"Updated workbook: {WORKBOOK_PATH}")
    print(f"Filled rows with Tokyo-only unique matches: {updated_rows}")
    print(f"Still blank due to Tokyo duplicates: {remaining_blank_duplicate_rows}")
    print(f"Still blank due to no Tokyo match: {remaining_blank_not_found_rows}")


if __name__ == "__main__":
    main()
