#!/usr/bin/env python3
"""Build a verified CSV for Tokyo travel businesses from registration 8074 backward.

Data sources:
- Uploaded Tokyo PDF list of registered travel businesses
- Official National Tax Agency corporate number bulk download (nationwide)

Rules implemented from the task:
- Read all merchants from registration number 8074 backward
- Fill only real information from the official NTA source
- If the same company name matches multiple corporations, leave result fields blank
- If nothing can be found, leave result fields blank
- Prefix the official address with the postal code
"""

from __future__ import annotations

import csv
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from pypdf import PdfReader


WORKSPACE = Path("/workspace")
PDF_PATH = Path(
    "/home/ubuntu/.cursor/projects/workspace/uploads/"
    "___________________________________________.pdf"
)
NTA_CSV_PATH = WORKSPACE / "data/raw/00_zenkoku_all_20260331.csv"
OUTPUT_PATH = WORKSPACE / "output/8074_backward_houjinbangou.csv"
SUCCESS_XLSX_PATH = WORKSPACE / "output/8074_backward_success_only.xlsx"

MAX_REGISTRATION_NUMBER = 8074
NTA_NAME_COL = 6
NTA_CORPORATE_NUMBER_COL = 1
NTA_PREF_COL = 9
NTA_CITY_COL = 10
NTA_ADDRESS_COL = 11
NTA_POSTCODE_COL = 15

ENTRY_RE = re.compile(r"^(\d+)\s+(第２種|第３種|地域限定|代理業|手配業)\s+(\S+)\s+(.*)$")


def normalize_name(text: str) -> str:
    """Normalize names conservatively for exact matching."""
    text = unicodedata.normalize("NFKC", text or "")
    return "".join(ch for ch in text if not ch.isspace())


def strip_trailing_alias(name: str) -> str:
    """Drop a trailing parenthesized trade name to search the legal name."""
    return re.sub(r"[\(（].*?[\)）]\s*$", "", name).strip()


def format_postcode(postcode: str) -> str:
    digits = re.sub(r"\D", "", postcode or "")
    if len(digits) == 7:
        return f"〒{digits[:3]}-{digits[3:]}"
    return ""


def format_address(postcode: str, prefecture: str, city: str, address: str) -> str:
    body = f"{prefecture or ''}{city or ''}{address or ''}".strip()
    formatted_postcode = format_postcode(postcode)
    if formatted_postcode and body:
        return f"{formatted_postcode} {body}"
    if body:
        return body
    return ""


def parse_pdf_entries() -> list[dict[str, str]]:
    reader = PdfReader(str(PDF_PATH))
    by_registration_number: dict[int, dict[str, str]] = {}

    for page_no, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        for raw_line in text.splitlines():
            line = raw_line.strip()
            match = ENTRY_RE.match(line)
            if not match:
                continue

            registration_number = int(match.group(1))
            if registration_number > MAX_REGISTRATION_NUMBER:
                continue

            remainder = match.group(4)
            association = ""
            for suffix in (" ＡＮＴＡ", " ＪＡＴＡ", " ANTA", " JATA"):
                if remainder.endswith(suffix):
                    association = suffix.strip()
                    remainder = remainder[: -len(suffix)]
                    break

            original_name = remainder.strip()
            by_registration_number[registration_number] = {
                "registration_number": str(registration_number),
                "business_type": match.group(2),
                "registration_date": match.group(3),
                "name_in_pdf": original_name,
                "search_name": strip_trailing_alias(original_name),
                "association": association,
                "source_page": str(page_no),
            }

    return [by_registration_number[key] for key in sorted(by_registration_number)]


def scan_nta_matches(search_names: set[str]) -> dict[str, list[dict[str, str]]]:
    matches: dict[str, list[dict[str, str]]] = defaultdict(list)

    with NTA_CSV_PATH.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if len(row) <= NTA_POSTCODE_COL:
                continue
            key = normalize_name(row[NTA_NAME_COL])
            if key not in search_names:
                continue
            matches[key].append(
                {
                    "corporate_number": row[NTA_CORPORATE_NUMBER_COL],
                    "official_name": row[NTA_NAME_COL],
                    "postcode": row[NTA_POSTCODE_COL],
                    "address": format_address(
                        row[NTA_POSTCODE_COL],
                        row[NTA_PREF_COL],
                        row[NTA_CITY_COL],
                        row[NTA_ADDRESS_COL],
                    ),
                }
            )

    return matches


def build_rows(entries: list[dict[str, str]]) -> list[dict[str, str]]:
    search_names = {normalize_name(entry["search_name"]) for entry in entries}
    matches = scan_nta_matches(search_names)

    rows: list[dict[str, str]] = []
    for entry in entries:
        key = normalize_name(entry["search_name"])
        candidates = matches.get(key, [])

        row = dict(entry)
        row["match_count"] = str(len(candidates))
        row["official_name"] = ""
        row["corporate_number"] = ""
        row["official_address_with_postcode"] = ""

        if len(candidates) == 1:
            row["match_status"] = "unique_match"
            row["official_name"] = candidates[0]["official_name"]
            row["corporate_number"] = candidates[0]["corporate_number"]
            row["official_address_with_postcode"] = candidates[0]["address"]
        elif len(candidates) == 0:
            row["match_status"] = "not_found"
        else:
            row["match_status"] = "duplicate_name"

        rows.append(row)

    return rows


def write_csv(rows: list[dict[str, str]]) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "registration_number",
        "business_type",
        "registration_date",
        "name_in_pdf",
        "search_name",
        "association",
        "source_page",
        "match_status",
        "match_count",
        "official_name",
        "corporate_number",
        "official_address_with_postcode",
    ]

    with OUTPUT_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_success_xlsx(rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "成功匹配"

    fieldnames = [
        "registration_number",
        "business_type",
        "registration_date",
        "name_in_pdf",
        "association",
        "official_name",
        "corporate_number",
        "official_address_with_postcode",
    ]
    headers = {
        "registration_number": "登録番号",
        "business_type": "業種",
        "registration_date": "登録年月日",
        "name_in_pdf": "旅行業者名（PDF）",
        "association": "正会員加入団体",
        "official_name": "法人名（国税庁）",
        "corporate_number": "法人番号",
        "official_address_with_postcode": "本所住所（郵便番号付き）",
    }

    for col_idx, key in enumerate(fieldnames, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=headers[key])
        cell.font = Font(bold=True)

    success_rows = [
        row for row in rows if row.get("match_status") == "unique_match"
    ]
    for row_idx, row in enumerate(success_rows, start=2):
        for col_idx, key in enumerate(fieldnames, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = min(max(max_length + 2, 12), 60)
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    workbook.save(SUCCESS_XLSX_PATH)


def main() -> None:
    if not PDF_PATH.exists():
        raise FileNotFoundError(f"PDF not found: {PDF_PATH}")
    if not NTA_CSV_PATH.exists():
        raise FileNotFoundError(f"NTA CSV not found: {NTA_CSV_PATH}")

    rows = build_rows(parse_pdf_entries())
    write_csv(rows)
    write_success_xlsx(rows)

    unique_count = sum(1 for row in rows if row["match_status"] == "unique_match")
    duplicate_count = sum(1 for row in rows if row["match_status"] == "duplicate_name")
    not_found_count = sum(1 for row in rows if row["match_status"] == "not_found")

    print(f"Wrote {len(rows)} rows to {OUTPUT_PATH}")
    print(f"Wrote {unique_count} successful rows to {SUCCESS_XLSX_PATH}")
    print(
        "Match summary:",
        f"unique={unique_count}",
        f"duplicate={duplicate_count}",
        f"not_found={not_found_count}",
    )


if __name__ == "__main__":
    main()
